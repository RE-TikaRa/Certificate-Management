import csv
import itertools
import json
import logging
from collections.abc import Callable, Iterable, Sequence
from dataclasses import dataclass
from datetime import date as py_date, datetime as py_datetime
from pathlib import Path
from typing import Any, cast

import pandas as pd
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from ..config import TEMPLATES_DIR
from ..data.database import Database
from ..data.models import Attachment, Award, AwardFlagValue, AwardMember, CustomFlag, ImportJob
from .attachment_manager import AttachmentManager
from .audit_logger import EntityType, get_audit_logger

logger = logging.getLogger(__name__)

TEMPLATE_HEADERS = [
    "比赛名称",
    "获奖日期",
    "赛事级别",
    "奖项等级",
    "证书编号",
    "备注",
    "成员",
    "附件路径",
]
RELATIVE_ATTACHMENT_HEADER = "附件相对路径"


@dataclass
class ImportResult:
    total: int
    success: int
    failed: int
    errors: list[str]
    error_file: Path | None = None


class ImportExportService:
    def __init__(self, db: Database, attachments: AttachmentManager, flags=None):
        self.db = db
        self.attachments = attachments
        self.flags = flags
        self._ensure_template()

    def get_awards_template_path(self, fmt: str = "xlsx") -> Path:
        """返回荣誉导入模板路径（csv/xlsx）。"""
        self._ensure_template()
        suffix = fmt.lower().lstrip(".")
        filename = "awards_template.xlsx" if suffix == "xlsx" else "awards_template.csv"
        return TEMPLATES_DIR / filename

    def _ensure_template(self) -> None:
        csv_path = TEMPLATES_DIR / "awards_template.csv"
        if not csv_path.exists():
            with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(TEMPLATE_HEADERS)
        xlsx_path = TEMPLATES_DIR / "awards_template.xlsx"
        if not xlsx_path.exists():
            df = pd.DataFrame(columns=pd.Index(TEMPLATE_HEADERS))
            df.to_excel(xlsx_path, index=False)

    def export_awards(self, export_path: Path, awards: Sequence[Award]) -> Path:
        flag_defs: list[CustomFlag] = []
        flag_values: dict[int, dict[str, bool]] = {}
        if self.flags:
            flag_defs = self.flags.list_flags(enabled_only=True)
            flag_values = self.flags.get_flags_for_awards([award.id for award in awards])

        award_ids = [award.id for award in awards]
        loaded_by_id: dict[int, Award] = {}
        attachment_by_award: dict[int, list[Attachment]] = {}
        if award_ids:
            with self.db.session_scope() as session:
                loaded = session.scalars(
                    select(Award)
                    .where(Award.id.in_(award_ids))
                    .options(selectinload(Award.award_members), selectinload(Award.attachments))
                ).all()
                loaded_by_id = {a.id: a for a in loaded}
                for a in loaded:
                    attachment_by_award[a.id] = [att for att in a.attachments if not att.deleted]

        attachments_root = self.attachments.ensure_root()
        export_path.parent.mkdir(parents=True, exist_ok=True)

        def build_row(src_award: Award) -> dict[str, object]:
            members = ",".join(getattr(src_award, "member_names", []) or [])
            attachment_items = attachment_by_award.get(src_award.id, [])
            attachment_paths = [
                str((attachments_root / att.relative_path).resolve())
                for att in attachment_items
                if isinstance(att.relative_path, str) and att.relative_path
            ]
            relative_paths = [
                str(att.relative_path)
                for att in attachment_items
                if isinstance(att.relative_path, str) and att.relative_path
            ]
            row: dict[str, object] = {
                "比赛名称": src_award.competition_name,
                "获奖日期": src_award.award_date.isoformat(),
                "赛事级别": src_award.level,
                "奖项等级": src_award.rank,
                "证书编号": src_award.certificate_code,
                "备注": src_award.remarks,
                "成员": members,
                "附件路径": ";".join(attachment_paths),
                RELATIVE_ATTACHMENT_HEADER: ";".join(relative_paths),
                "附件数量": len(attachment_paths),
            }
            if flag_defs:
                values = flag_values.get(src_award.id, {})
                for flag in flag_defs:
                    col = f"{flag.label} ({flag.key})"
                    row[col] = int(values.get(flag.key, flag.default_value))
            return row

        if export_path.suffix.lower() == ".csv":
            flag_headers = [f"{flag.label} ({flag.key})" for flag in flag_defs]
            headers = [*TEMPLATE_HEADERS, RELATIVE_ATTACHMENT_HEADER, "附件数量", *flag_headers]
            with export_path.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
                writer.writeheader()
                for award in awards:
                    src_award = loaded_by_id.get(award.id, award)
                    writer.writerow(build_row(src_award))
        else:
            rows = []
            for award in awards:
                src_award = loaded_by_id.get(award.id, award)
                rows.append(build_row(src_award))
            df = pd.DataFrame(rows)
            df.to_excel(export_path, index=False)
        try:
            size_mb = export_path.stat().st_size / (1024 * 1024) if export_path.exists() else 0
        except Exception:
            size_mb = 0
        audit = get_audit_logger()
        audit.log_export(str(export_path), EntityType.AWARD, len(awards), size_mb)
        logger.info("Exported %s awards to %s", len(awards), export_path)
        return export_path

    def import_from_file(
        self,
        file_path: Path,
        *,
        progress_callback: Callable[[int, int, float], None] | None = None,
        dry_run: bool = False,
    ) -> ImportResult:
        flag_defs: list[CustomFlag] = self.flags.list_flags(enabled_only=True) if self.flags else []
        frame_iter: Iterable[pd.DataFrame] | None = None
        row_iter: Iterable[object] | None = None
        total = 0
        columns: list[str] = []
        encoding = "utf-8-sig"
        try:
            if file_path.suffix.lower() == ".xlsx":
                row_iter, total, columns = self._iter_xlsx_rows(file_path)
            else:
                try:
                    reader = pd.read_csv(file_path, encoding="utf-8-sig", chunksize=500)
                    first_chunk = next(reader, None)
                    if first_chunk is None:
                        frame_iter = None
                    else:
                        frame_iter = itertools.chain([first_chunk], reader)
                        columns = list(first_chunk.columns)
                    encoding = "utf-8-sig"
                except Exception:
                    reader = pd.read_csv(file_path, encoding="gbk", chunksize=500)
                    first_chunk = next(reader, None)
                    if first_chunk is None:
                        frame_iter = None
                    else:
                        frame_iter = itertools.chain([first_chunk], reader)
                        columns = list(first_chunk.columns)
                    encoding = "gbk"
                if frame_iter is not None:

                    def _iter_rows() -> Iterable[object]:
                        for frame in frame_iter:
                            for _, row in frame.iterrows():
                                yield row

                    row_iter = _iter_rows()
        except Exception as exc:
            return ImportResult(total=0, success=0, failed=0, errors=[str(exc)])
        if row_iter is None or not columns:
            return ImportResult(total=0, success=0, failed=0, errors=[])

        def count_csv_rows() -> int:
            if file_path.suffix.lower() == ".xlsx":
                return total
            try:
                with file_path.open("r", encoding=encoding, errors="replace", newline="") as f:
                    row_count = max(sum(1 for _ in f) - 1, 0)
            except Exception:
                return total
            return row_count

        required_cols = {"比赛名称", "获奖日期", "赛事级别", "奖项等级"}
        if not required_cols.issubset(columns):
            missing = ", ".join(required_cols - set(columns))
            return ImportResult(total=0, success=0, failed=0, errors=[f"缺少列: {missing}"])

        total = count_csv_rows()
        success = 0
        errors: list[str] = []
        error_rows: list[dict] = []

        import time

        start_time = time.time()

        # flag 列映射：优先匹配 "label (key)"，其次 label
        flag_col_map: dict[str, str] = {}
        for flag in flag_defs:
            preferred = f"{flag.label} ({flag.key})"
            if preferred in columns:
                flag_col_map[flag.key] = preferred
            elif flag.label in columns:
                flag_col_map[flag.key] = flag.label
            else:
                flag_col_map[flag.key] = ""  # 缺失列，用默认值

        def clean_cell(value) -> str:
            if value is None:
                return ""
            if isinstance(value, float) and pd.isna(value):
                return ""
            if pd.isna(value):
                return ""
            text = str(value).strip()
            return "" if text.lower() == "nan" else text

        def get_cell(row, key: str):
            if hasattr(row, "get"):
                try:
                    return row.get(key, None)
                except Exception:
                    return None
            try:
                return row[key]
            except Exception:
                return None

        def row_to_dict(row) -> dict:
            if hasattr(row, "to_dict"):
                return row.to_dict()
            if isinstance(row, dict):
                return dict(row)
            return {}

        def handle_row(session: Session, row_index: int, row) -> None:
            nonlocal success
            raw_date = get_cell(row, "获奖日期")
            timestamp: object
            if isinstance(raw_date, pd.Timestamp):
                timestamp = raw_date
            elif isinstance(raw_date, py_datetime):
                timestamp = pd.Timestamp(raw_date)
            elif isinstance(raw_date, py_date):
                timestamp = pd.Timestamp(py_datetime(raw_date.year, raw_date.month, raw_date.day))
            elif isinstance(raw_date, (int, float)) and not pd.isna(raw_date):
                # Excel serial date (days since 1899-12-30 in pandas)
                ts = pd.to_datetime(raw_date, unit="D", origin="1899-12-30")
                if pd.isna(ts):
                    raise ValueError("获奖日期无法解析")
                timestamp = ts
            else:
                ts = pd.to_datetime(clean_cell(raw_date))
                if pd.isna(ts):
                    raise ValueError("获奖日期无法解析")
                timestamp = ts
            if cast(bool, pd.isna(timestamp)):
                raise ValueError("获奖日期无法解析")
            award = Award(
                competition_name=clean_cell(get_cell(row, "比赛名称")),
                award_date=cast(pd.Timestamp, timestamp).date(),
                level=clean_cell(get_cell(row, "赛事级别")),
                rank=clean_cell(get_cell(row, "奖项等级")),
                certificate_code=clean_cell(get_cell(row, "证书编号") or "") or None,
                remarks=clean_cell(get_cell(row, "备注") or "") or None,
            )
            session.add(award)
            session.flush()

            members = self._parse_items(clean_cell(get_cell(row, "成员") or ""))
            award.award_members = [
                AwardMember(member_name=name, sort_order=index) for index, name in enumerate(members)
            ]
            session.flush()

            if flag_defs and not dry_run:
                values: list[AwardFlagValue] = []
                for flag in flag_defs:
                    col = flag_col_map.get(flag.key, "")
                    raw = get_cell(row, col) if col else None
                    value = self._parse_flag_value(raw, default=flag.default_value)
                    values.append(AwardFlagValue(award_id=award.id, flag_key=flag.key, value=value))
                session.add_all(values)

            attachment_paths = self._parse_items(clean_cell(get_cell(row, "附件路径") or ""), sep=";")
            relative_paths = self._parse_items(clean_cell(get_cell(row, RELATIVE_ATTACHMENT_HEADER) or ""), sep=";")
            files = self._resolve_attachment_paths(attachment_paths, relative_paths)
            if files and not dry_run:
                self.attachments.save_attachments(award.id, award.competition_name, files, session=session)

            if not dry_run:
                self.db.refresh_award_fts(award.id, session=session)

            success += 1

        def process_rows(session: Session) -> None:
            for row_index, row in enumerate(row_iter or []):
                try:
                    with session.begin_nested():
                        handle_row(session, row_index, row)
                except Exception as exc:
                    errors.append(f"第 {row_index + 2} 行: {exc}")
                    error_rows.append({"行号": row_index + 2, "索引": row_index, "错误": str(exc), **row_to_dict(row)})

                current = row_index + 1
                if progress_callback and current % 10 == 0:
                    elapsed = time.time() - start_time
                    rate = elapsed / max(current, 1)
                    remaining = max(total - current, 0) * rate
                    progress_callback(current, total, float(remaining))

        if dry_run:
            # Strict dry-run: no DB writes, no attachment copy, no ImportJob, no error file.
            with self.db.engine.connect() as connection:
                transaction = connection.begin()
                session = Session(bind=connection, expire_on_commit=False)
                try:
                    process_rows(session)
                finally:
                    session.close()
                    transaction.rollback()
        else:
            with self.db.session_scope() as session:
                process_rows(session)
                status = "success" if not errors else ("failed" if success <= 0 else "partial")
                session.add(
                    ImportJob(
                        filename=file_path.name,
                        status=status,
                        message="\n".join(errors) if errors else None,
                    )
                )

        error_file: Path | None = None
        if error_rows and not dry_run:
            error_file = file_path.parent / f"{file_path.stem}_errors.csv"
            try:
                pd.DataFrame(error_rows).to_csv(error_file, index=False, encoding="utf-8-sig")
            except Exception:
                error_file = None

        result = ImportResult(
            total=total, success=success, failed=total - success, errors=errors, error_file=error_file
        )
        if not dry_run:
            audit = get_audit_logger()
            audit.log_import(str(file_path), result.total, result.success, result.failed, result.errors)
        return result

    def _parse_items(self, value: str, sep: str = ",") -> list[str]:
        unique: list[str] = []
        seen: set[str] = set()
        for item in value.split(sep):
            cleaned = item.strip()
            if cleaned and cleaned.lower() not in seen:
                seen.add(cleaned.lower())
                unique.append(cleaned)
        return unique

    def _resolve_attachment_paths(self, absolute_paths: list[str], relative_paths: list[str]) -> list[Path]:
        root = self.attachments.ensure_root()
        resolved: list[Path] = []
        seen: set[str] = set()

        def add_path(path: Path) -> None:
            key = str(path.resolve()).lower() if path.exists() else str(path).lower()
            if key in seen:
                return
            seen.add(key)
            resolved.append(path)

        for raw in absolute_paths:
            if not raw:
                continue
            candidate = Path(raw)
            if candidate.exists():
                add_path(candidate)
                continue
            if not candidate.is_absolute():
                fallback = root / candidate
                if fallback.exists():
                    add_path(fallback)
                    continue
            if candidate.is_absolute():
                rel_candidate = Path(*candidate.parts[1:]) if candidate.drive else Path(str(candidate).lstrip("/\\"))
                fallback = root / rel_candidate
                if fallback.exists():
                    add_path(fallback)

        for raw in relative_paths:
            if not raw:
                continue
            candidate = root / Path(raw)
            if candidate.exists():
                add_path(candidate)

        return resolved

    def _parse_flag_value(self, value, default: bool) -> bool:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return bool(default)
        if isinstance(value, (int, bool)):
            return bool(value)
        text = str(value).strip().lower()
        if text in {"1", "true", "t", "yes", "y", "是", "对", "勾", "勾选"}:
            return True
        if text in {"0", "false", "f", "no", "n", "否", "不", "未"}:
            return False
        return bool(default)

    def _iter_xlsx_rows(self, file_path: Path) -> tuple[Iterable[dict[str, Any]], int, list[str]]:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValueError("缺少 Excel 解析依赖：openpyxl（请先执行 uv sync 安装依赖）") from exc

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        if sheet is None:
            workbook.close()
            return iter([]), 0, []
        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            workbook.close()
            return iter([]), 0, []

        headers = [str(value).strip() if value is not None else "" for value in header_row]
        total = max(int(sheet.max_row or 0) - 1, 0)

        def _iter() -> Iterable[dict[str, Any]]:
            try:
                for row in rows:
                    if row is None:
                        continue
                    data: dict[str, Any] = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
                    yield data
            finally:
                workbook.close()

        return _iter(), total, headers

    def list_jobs(self, limit: int = 20) -> list[ImportJob]:
        with self.db.session_scope() as session:
            q = select(ImportJob).order_by(ImportJob.created_at.desc()).limit(max(1, limit))
            return list(session.scalars(q).all())

    def export_award_pdf(self, award_id: int, output_path: Path) -> Path:
        with self.db.session_scope() as session:
            award = session.scalar(select(Award).where(Award.id == award_id).options(selectinload(Award.award_members)))
            if award is None:
                raise ValueError("未找到要导出的荣誉记录")

            output_path.parent.mkdir(parents=True, exist_ok=True)
            try:
                import fitz
            except Exception as exc:
                raise ValueError("缺少 PDF 生成依赖：PyMuPDF（请先执行 uv sync 安装依赖）") from exc

            doc = fitz.open()
            page = doc.new_page(width=595, height=842)

            title_rect = fitz.Rect(0, 50, page.rect.width, 90)
            page.insert_textbox(title_rect, "荣誉证书", fontsize=24, fontname="helv", align=1)

            y = 120
            line_gap = 24

            def add_line(label: str, value: str | None) -> None:
                nonlocal y
                text = value or "—"
                page.insert_text((60, y), f"{label}：{text}", fontsize=12, fontname="helv")
                y += line_gap

            add_line("比赛名称", award.competition_name)
            add_line("获奖日期", award.award_date.isoformat())
            add_line("赛事级别", award.level)
            add_line("奖项等级", award.rank)
            add_line("证书编号", award.certificate_code)

            members = "、".join(award.member_names)
            members_rect = fitz.Rect(60, y, page.rect.width - 60, y + 80)
            page.insert_textbox(members_rect, f"成员：{members or '—'}", fontsize=12, fontname="helv")
            y = members_rect.y1 + 6

            if award.remarks:
                remarks_rect = fitz.Rect(60, y, page.rect.width - 60, y + 80)
                page.insert_textbox(remarks_rect, f"备注：{award.remarks}", fontsize=11, fontname="helv")

            qr_payload = self._build_qr_payload(award)
            qr_matrix = self._build_qr_matrix(qr_payload)
            qr_rect = fitz.Rect(
                page.rect.width - 180, page.rect.height - 180, page.rect.width - 40, page.rect.height - 40
            )
            self._draw_qr_to_page(page, qr_matrix, qr_rect)
            page.insert_text((qr_rect.x0, qr_rect.y0 - 10), "证书二维码", fontsize=9, fontname="helv")

            doc.save(output_path)
            doc.close()

            try:
                size_mb = output_path.stat().st_size / (1024 * 1024)
            except Exception:
                size_mb = 0
            audit = get_audit_logger()
            audit.log_export(str(output_path), EntityType.AWARD, 1, size_mb)

        return output_path

    def export_award_qr(self, award_id: int, output_path: Path) -> Path:
        with self.db.session_scope() as session:
            award = session.scalar(select(Award).where(Award.id == award_id).options(selectinload(Award.award_members)))
            if award is None:
                raise ValueError("未找到要导出的荣誉记录")

            output_path.parent.mkdir(parents=True, exist_ok=True)
            matrix = self._build_qr_matrix(self._build_qr_payload(award))
            self._write_qr_png(matrix, output_path)
            try:
                size_mb = output_path.stat().st_size / (1024 * 1024)
            except Exception:
                size_mb = 0
            audit = get_audit_logger()
            audit.log_export(str(output_path), EntityType.AWARD, 1, size_mb)
            return output_path

    def _build_qr_payload(self, award: Award) -> str:
        payload = {
            "id": award.id,
            "name": award.competition_name,
            "date": award.award_date.isoformat(),
            "level": award.level,
            "rank": award.rank,
            "code": award.certificate_code,
        }
        return json.dumps(payload, ensure_ascii=False)

    def _build_qr_matrix(self, data: str) -> list[list[bool]]:
        try:
            import importlib

            qrcode = importlib.import_module("qrcode")
        except Exception as exc:
            raise ValueError("缺少二维码依赖：qrcode（请先执行 uv sync 安装依赖）") from exc
        qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=1, border=2)
        qr.add_data(data)
        qr.make(fit=True)
        matrix = qr.get_matrix()
        if not matrix:
            raise ValueError("二维码生成失败")
        return matrix

    def _draw_qr_to_page(self, page, matrix: list[list[bool]], rect) -> None:
        try:
            import fitz
        except Exception:
            return
        size = len(matrix)
        if size <= 0:
            return
        page.draw_rect(rect, color=(1, 1, 1), fill=(1, 1, 1))
        module = min(rect.width, rect.height) / size
        for row_idx, row in enumerate(matrix):
            for col_idx, value in enumerate(row):
                if not value:
                    continue
                x0 = rect.x0 + col_idx * module
                y0 = rect.y0 + row_idx * module
                page.draw_rect(fitz.Rect(x0, y0, x0 + module, y0 + module), color=(0, 0, 0), fill=(0, 0, 0))

    def _write_qr_png(self, matrix: list[list[bool]], output_path: Path) -> None:
        from PySide6.QtCore import Qt
        from PySide6.QtGui import QColor, QImage, QPainter

        size = len(matrix)
        if size <= 0:
            raise ValueError("二维码内容为空")
        module = 6
        img_size = size * module
        image = QImage(img_size, img_size, QImage.Format.Format_ARGB32)
        image.fill(QColor("white"))

        painter = QPainter(image)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor("black"))
        for row_idx, row in enumerate(matrix):
            for col_idx, value in enumerate(row):
                if value:
                    painter.drawRect(col_idx * module, row_idx * module, module, module)
        painter.end()

        if not image.save(str(output_path)):
            raise ValueError("二维码保存失败")
