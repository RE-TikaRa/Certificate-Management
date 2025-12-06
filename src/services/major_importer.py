"""
Excel 导入专业数据的辅助工具。
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from .major_service import MajorService


def read_majors_from_excel(excel_path: Path) -> list[str]:
    """
    从 Excel 文件中读取专业名称列表。

    Args:
        excel_path: Excel 文件路径

    Returns:
        专业名称列表
    """
    from openpyxl import load_workbook

    workbook = load_workbook(excel_path)
    majors: list[str] = []
    try:
        worksheet = workbook.active
        if worksheet is None:
            return majors

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if len(row or ()) > 2:
                name = row[2]
                if name:
                    text = str(name).strip()
                    if text and text != "专业名称":
                        majors.append(text)
    finally:
        workbook.close()

    return majors


def import_majors_from_excel(service: MajorService, excel_path: Path) -> int:
    """
    使用 MajorService 将 Excel 中的专业批量导入数据库。

    Args:
        service: 专业服务实例
        excel_path: Excel 文件路径

    Returns:
        成功导入的数量
    """
    majors = read_majors_from_excel(excel_path)
    if not majors:
        return 0
    return service.replace_all_majors(majors)
