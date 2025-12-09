"""
读取学校、专业目录以及学校-专业映射的数据文件。
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from .academic_types import MajorCatalogInput, SchoolInput, SchoolMajorMappingInput


def read_school_list(csv_path: Path) -> list[SchoolInput]:
    """从 CSV 获取学校名称与代码"""
    schools: list[SchoolInput] = []
    with csv_path.open(encoding="utf-8-sig") as fp:
        reader = csv.DictReader(fp)
        for row in reader:
            name = (row.get("学校名称") or "").strip()
            code = (row.get("学校标识码") or row.get("学校代码") or "").strip() or None
            region = (row.get("所在地") or "").strip() or None
            if name:
                schools.append(SchoolInput(name=name, code=code, region=region))
    return schools


def read_major_catalog(csv_path: Path) -> list[MajorCatalogInput]:
    records: list[MajorCatalogInput] = []
    with csv_path.open(encoding="utf-8-sig") as fp:
        reader = csv.DictReader(fp)
        for row in reader:
            major_name = (row.get("major_name") or row.get("专业名称") or "").strip()
            major_code = (row.get("major_code") or row.get("专业代码") or "").strip() or None
            if not major_name:
                continue
            records.append(
                MajorCatalogInput(
                    major_name=major_name,
                    major_code=major_code,
                    discipline_code=(row.get("discipline_code") or row.get("学科门类码") or "").strip() or None,
                    discipline_name=(row.get("discipline_name") or row.get("学科门类") or "").strip() or None,
                    class_code=(row.get("class_code") or row.get("专业类代码") or "").strip() or None,
                    class_name=(row.get("class_name") or row.get("专业类") or "").strip() or None,
                    category=(row.get("category") or row.get("科类") or "").strip() or None,
                )
            )
    return records


def read_school_major_mappings(excel_path: Path) -> list[SchoolMajorMappingInput]:
    """解析学校-专业-学院映射"""
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    records: list[SchoolMajorMappingInput] = []
    try:
        sheet = workbook.active
        if sheet is None:
            return records

        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return records

        header_map = {str(value).strip(): idx for idx, value in enumerate(header) if value}

        def get(row: tuple, keys: tuple[str, ...]) -> str | None:
            for key in keys:
                idx = header_map.get(key)
                if idx is not None and idx < len(row):
                    value = row[idx]
                    if value is not None:
                        text = str(value).strip()
                        if text:
                            return text
            return None

        for values in sheet.iter_rows(min_row=2, values_only=True):
            school_name = get(values, ("学校", "学校名称"))
            major_name = get(values, ("专业名称", "专业"))
            if not school_name or not major_name:
                continue
            records.append(
                SchoolMajorMappingInput(
                    school_name=school_name,
                    school_code=get(values, ("学校代码", "学校标识码")),
                    major_code=get(values, ("专业代码",)),
                    major_name=major_name,
                    category=get(values, ("科类", "专业类")),
                    college_name=get(values, ("学院", "院系")),
                )
            )
    finally:
        workbook.close()
    return records
